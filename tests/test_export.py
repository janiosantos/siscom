"""
Testes para módulo de exportação de dados
"""
import pytest
import io
from datetime import date, datetime, timedelta
from decimal import Decimal

from app.modules.export.service import ExportService
from app.modules.export.schemas import ExportFiltros
from app.modules.vendas.models import Venda
from app.modules.produtos.models import Produto
from app.modules.categorias.models import Categoria
from app.modules.clientes.models import Cliente
from app.modules.auth.models import User


@pytest.mark.asyncio
async def test_export_dashboard_stats_csv(db_session, sample_vendas):
    """Testa export de estatísticas do dashboard em CSV"""
    service = ExportService(db_session)

    result = await service.export_dashboard_stats(
        formato="csv",
        tipo="stats",
        filtros=None
    )

    assert isinstance(result, io.BytesIO)
    content = result.getvalue().decode('utf-8-sig')
    assert "Métrica" in content
    assert "Valor" in content
    assert "Vendas" in content


@pytest.mark.asyncio
async def test_export_dashboard_vendas_por_dia_excel(db_session, sample_vendas):
    """Testa export de vendas por dia em Excel"""
    try:
        import openpyxl
        OPENPYXL_AVAILABLE = True
    except ImportError:
        OPENPYXL_AVAILABLE = False

    if not OPENPYXL_AVAILABLE:
        pytest.skip("openpyxl não disponível")

    service = ExportService(db_session)

    hoje = date.today()
    filtros = ExportFiltros(
        data_inicio=hoje - timedelta(days=7),
        data_fim=hoje
    )

    result = await service.export_dashboard_stats(
        formato="excel",
        tipo="vendas_dia",
        filtros=filtros
    )

    assert isinstance(result, io.BytesIO)

    # Validar que é um arquivo Excel válido
    wb = openpyxl.load_workbook(result)
    ws = wb.active
    assert ws.cell(1, 1).value == "Vendas por Dia"  # Título
    assert ws.cell(2, 1).value == "Data"  # Header


@pytest.mark.asyncio
async def test_export_produtos_mais_vendidos(db_session, sample_vendas):
    """Testa export de produtos mais vendidos"""
    service = ExportService(db_session)

    result = await service.export_dashboard_stats(
        formato="csv",
        tipo="produtos",
        filtros=None
    )

    assert isinstance(result, io.BytesIO)
    content = result.getvalue().decode('utf-8-sig')
    assert "Produto" in content
    assert "Quantidade" in content
    assert "Faturamento" in content


@pytest.mark.asyncio
async def test_export_vendas_por_vendedor(db_session, sample_vendas):
    """Testa export de vendas por vendedor"""
    service = ExportService(db_session)

    result = await service.export_dashboard_stats(
        formato="csv",
        tipo="vendedores",
        filtros=None
    )

    assert isinstance(result, io.BytesIO)
    content = result.getvalue().decode('utf-8-sig')
    assert "Vendedor" in content
    assert "Vendas" in content
    assert "Ticket Médio" in content


@pytest.mark.asyncio
async def test_export_vendas_por_status(db_session, sample_vendas):
    """Testa export de vendas por status"""
    service = ExportService(db_session)

    result = await service.export_dashboard_stats(
        formato="csv",
        tipo="status",
        filtros=None
    )

    assert isinstance(result, io.BytesIO)
    content = result.getvalue().decode('utf-8-sig')
    assert "Status" in content
    assert "Quantidade" in content
    assert "Percentual" in content


@pytest.mark.asyncio
async def test_export_orcamentos(db_session, sample_orcamentos):
    """Testa export de orçamentos"""
    service = ExportService(db_session)

    result = await service.export_orcamentos(
        formato="csv",
        filtros=None,
        ids=None
    )

    assert isinstance(result, io.BytesIO)
    content = result.getvalue().decode('utf-8-sig')
    assert "ID" in content
    assert "Data" in content
    assert "Cliente" in content
    assert "Valor Total" in content


@pytest.mark.asyncio
async def test_export_orcamentos_filtro_data(db_session, sample_orcamentos):
    """Testa export de orçamentos com filtro de data"""
    service = ExportService(db_session)

    hoje = date.today()
    filtros = ExportFiltros(
        data_inicio=hoje - timedelta(days=30),
        data_fim=hoje
    )

    result = await service.export_orcamentos(
        formato="csv",
        filtros=filtros,
        ids=None
    )

    assert isinstance(result, io.BytesIO)


@pytest.mark.asyncio
async def test_export_orcamentos_por_ids(db_session, sample_orcamentos):
    """Testa export de orçamentos específicos por IDs"""
    service = ExportService(db_session)

    # Pegar IDs dos orçamentos de teste
    orcamento_ids = [orc.id for orc in sample_orcamentos[:2]]

    result = await service.export_orcamentos(
        formato="csv",
        filtros=None,
        ids=orcamento_ids
    )

    assert isinstance(result, io.BytesIO)
    content = result.getvalue().decode('utf-8-sig')

    # Deve conter exatamente 2 orçamentos + header
    lines = content.strip().split('\n')
    assert len(lines) == 3  # 1 header + 2 orçamentos


@pytest.mark.asyncio
async def test_export_vendas(db_session, sample_vendas):
    """Testa export de vendas"""
    service = ExportService(db_session)

    result = await service.export_vendas(
        formato="csv",
        filtros=None,
        ids=None
    )

    assert isinstance(result, io.BytesIO)
    content = result.getvalue().decode('utf-8-sig')
    assert "ID" in content
    assert "Data" in content
    assert "Cliente" in content
    assert "Status" in content
    assert "Forma Pagamento" in content
    assert "Valor Total" in content


@pytest.mark.asyncio
async def test_export_vendas_filtro_vendedor(db_session, sample_vendas, sample_user):
    """Testa export de vendas com filtro por vendedor"""
    service = ExportService(db_session)

    filtros = ExportFiltros(
        vendedor_id=sample_user.id
    )

    result = await service.export_vendas(
        formato="csv",
        filtros=filtros,
        ids=None
    )

    assert isinstance(result, io.BytesIO)


@pytest.mark.asyncio
async def test_export_produtos(db_session, sample_produtos):
    """Testa export de produtos"""
    service = ExportService(db_session)

    result = await service.export_produtos(
        formato="csv",
        categoria_id=None,
        apenas_ativos=True
    )

    assert isinstance(result, io.BytesIO)
    content = result.getvalue().decode('utf-8-sig')
    assert "Código" in content
    assert "Descrição" in content
    assert "Categoria" in content
    assert "Preço Venda" in content
    assert "Estoque" in content


@pytest.mark.asyncio
async def test_export_produtos_por_categoria(db_session, sample_produtos, sample_categoria):
    """Testa export de produtos filtrados por categoria"""
    service = ExportService(db_session)

    result = await service.export_produtos(
        formato="csv",
        categoria_id=sample_categoria.id,
        apenas_ativos=True
    )

    assert isinstance(result, io.BytesIO)


@pytest.mark.asyncio
async def test_export_produtos_todos(db_session, sample_produtos):
    """Testa export de todos os produtos (ativos e inativos)"""
    service = ExportService(db_session)

    result = await service.export_produtos(
        formato="csv",
        categoria_id=None,
        apenas_ativos=False
    )

    assert isinstance(result, io.BytesIO)


@pytest.mark.asyncio
async def test_export_excel_without_openpyxl(db_session, monkeypatch):
    """Testa que erro apropriado é lançado quando openpyxl não está disponível"""
    service = ExportService(db_session)

    # Simular openpyxl não disponível
    import app.modules.export.service as export_module
    monkeypatch.setattr(export_module, "OPENPYXL_AVAILABLE", False)

    with pytest.raises(ImportError, match="openpyxl não está instalado"):
        await service.export_dashboard_stats(
            formato="excel",
            tipo="stats",
            filtros=None
        )


# ===== Fixtures =====

@pytest.fixture
async def sample_categoria(db_session):
    """Cria categoria de teste"""
    categoria = Categoria(
        nome="Materiais",
        descricao="Materiais de Construção"
    )
    db_session.add(categoria)
    await db_session.commit()
    await db_session.refresh(categoria)
    return categoria


@pytest.fixture
async def sample_user(db_session):
    """Cria usuário de teste"""
    user = User(
        nome="Vendedor Teste",
        email="vendedor@teste.com",
        senha_hash="hash123",
        ativo=True
    )
    db_session.add(user)
    await db_session.commit()
    await db_session.refresh(user)
    return user


@pytest.fixture
async def sample_cliente(db_session):
    """Cria cliente de teste"""
    cliente = Cliente(
        nome="Cliente Teste",
        cpf_cnpj="12345678900",
        tipo_pessoa="F",
        ativo=True
    )
    db_session.add(cliente)
    await db_session.commit()
    await db_session.refresh(cliente)
    return cliente


@pytest.fixture
async def sample_produtos(db_session, sample_categoria):
    """Cria produtos de teste"""
    produtos = []
    for i in range(5):
        produto = Produto(
            codigo=f"PROD-{i:03d}",
            descricao=f"Produto Teste {i}",
            categoria_id=sample_categoria.id,
            unidade="UN",
            preco_custo=Decimal("50.00"),
            preco_venda=Decimal("100.00"),
            estoque_atual=Decimal("10.0"),
            ativo=True
        )
        db_session.add(produto)
        produtos.append(produto)

    await db_session.commit()
    for p in produtos:
        await db_session.refresh(p)

    return produtos


@pytest.fixture
async def sample_vendas(db_session, sample_user, sample_cliente, sample_produtos):
    """Cria vendas de teste"""
    vendas = []
    for i in range(3):
        venda = Venda(
            data_venda=datetime.now() - timedelta(days=i),
            cliente_id=sample_cliente.id,
            vendedor_id=sample_user.id,
            status="finalizada",
            forma_pagamento="dinheiro",
            valor_total=Decimal("300.00"),
            desconto=Decimal("0.00"),
            valor_final=Decimal("300.00")
        )
        db_session.add(venda)
        vendas.append(venda)

    await db_session.commit()
    for v in vendas:
        await db_session.refresh(v)

    return vendas


@pytest.fixture
async def sample_orcamentos(db_session, sample_user, sample_cliente):
    """Cria orçamentos de teste"""
    from app.modules.orcamentos.models import Orcamento

    orcamentos = []
    for i in range(3):
        orcamento = Orcamento(
            data_orcamento=datetime.now() - timedelta(days=i),
            cliente_id=sample_cliente.id,
            vendedor_id=sample_user.id,
            status="pendente",
            valor_total=Decimal("500.00"),
            desconto=Decimal("0.00"),
            validade=datetime.now() + timedelta(days=7)
        )
        db_session.add(orcamento)
        orcamentos.append(orcamento)

    await db_session.commit()
    for o in orcamentos:
        await db_session.refresh(o)

    return orcamentos
