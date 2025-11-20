"""
Service de Import/Export com suporte a múltiplos formatos
"""
from sqlalchemy.ext.asyncio import AsyncSession
from typing import List, Optional, Dict, Any, BinaryIO
from datetime import datetime
import csv
import json
import io
import tempfile
import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .repository import ImportExportRepository
from .models import ImportStatus, ExportStatus, ImportFormat, ExportFormat
from .schemas import (
    ImportPreviewResponse,
    ImportPreviewRow,
    ImportValidationError,
    ImportStatusResponse,
    ExportStatusResponse,
    ImportStartRequest,
    ExportRequest,
    ImportTemplateCreate,
    ImportTemplateResponse,
    BulkOperationResult,
    ImportExportStats
)
from app.core.exceptions import NotFoundException, BusinessException
from app.core.logging import get_logger

logger = get_logger(__name__)


class ImportExportService:
    """Service de Import/Export"""

    def __init__(self, db: AsyncSession):
        self.repository = ImportExportRepository(db)
        self.db = db

    # ========================================================================
    # IMPORT - PREVIEW E VALIDAÇÃO
    # ========================================================================

    async def preview_import(
        self,
        file: BinaryIO,
        format: ImportFormat,
        module: str,
        sample_size: int = 10
    ) -> ImportPreviewResponse:
        """
        Preview de importação com validação básica
        """
        logger.info(f"Gerando preview de importação: module={module}, format={format}")

        try:
            # Ler dados do arquivo
            if format == ImportFormat.CSV:
                data = self._read_csv(file)
            elif format == ImportFormat.EXCEL:
                data = self._read_excel(file)
            elif format == ImportFormat.JSON:
                data = self._read_json(file)
            else:
                raise BusinessException(f"Formato não suportado: {format}")

            if not data or len(data) == 0:
                raise BusinessException("Arquivo vazio ou inválido")

            # Extrair colunas
            columns = list(data[0].keys()) if data else []

            # Validar estrutura básica
            validation_errors = []
            sample_rows = []

            for idx, row in enumerate(data[:sample_size], start=1):
                errors = []

                # Validações básicas
                if not row or all(v is None or str(v).strip() == "" for v in row.values()):
                    errors.append("Linha vazia")

                # Validar tipos de dados básicos
                for col, value in row.items():
                    if value is not None and str(value).strip():
                        # Aqui poderiam ser adicionadas validações específicas por módulo
                        pass

                is_valid = len(errors) == 0

                sample_rows.append(ImportPreviewRow(
                    row_number=idx,
                    data=row,
                    is_valid=is_valid,
                    errors=errors
                ))

                if not is_valid:
                    for error in errors:
                        validation_errors.append(ImportValidationError(
                            row=idx,
                            column="*",
                            value="",
                            error=error
                        ))

            # Sugerir mapeamento de colunas
            mapping_suggestions = self._suggest_column_mapping(columns, module)

            return ImportPreviewResponse(
                total_rows=len(data),
                sample_rows=sample_rows,
                columns=columns,
                detected_format=format,
                validation_errors=validation_errors,
                is_valid=len(validation_errors) == 0,
                mapping_suggestions=mapping_suggestions
            )

        except Exception as e:
            logger.error(f"Erro ao gerar preview: {str(e)}")
            raise BusinessException(f"Erro ao processar arquivo: {str(e)}")

    async def validate_import(
        self,
        file: BinaryIO,
        format: ImportFormat,
        module: str,
        mapping: Dict[str, str]
    ) -> List[ImportValidationError]:
        """
        Validação completa da importação
        """
        logger.info(f"Validando importação: module={module}")

        errors = []

        try:
            # Ler dados
            if format == ImportFormat.CSV:
                data = self._read_csv(file)
            elif format == ImportFormat.EXCEL:
                data = self._read_excel(file)
            elif format == ImportFormat.JSON:
                data = self._read_json(file)
            else:
                raise BusinessException(f"Formato não suportado: {format}")

            # Validar cada linha
            for idx, row in enumerate(data, start=1):
                # Aplicar mapeamento
                mapped_row = {}
                for file_col, system_field in mapping.items():
                    if file_col in row:
                        mapped_row[system_field] = row[file_col]

                # Validar campos obrigatórios
                required_fields = self._get_required_fields(module)
                for field in required_fields:
                    if field not in mapped_row or not mapped_row[field]:
                        errors.append(ImportValidationError(
                            row=idx,
                            column=field,
                            value=mapped_row.get(field),
                            error=f"Campo obrigatório '{field}' não fornecido"
                        ))

                # Validações específicas por módulo
                module_errors = await self._validate_module_data(module, mapped_row, idx)
                errors.extend(module_errors)

            return errors

        except Exception as e:
            logger.error(f"Erro na validação: {str(e)}")
            raise BusinessException(f"Erro ao validar arquivo: {str(e)}")

    # ========================================================================
    # IMPORT - EXECUÇÃO
    # ========================================================================

    async def start_import(
        self,
        file: BinaryIO,
        filename: str,
        request: ImportStartRequest,
        user_id: Optional[int] = None
    ) -> ImportStatusResponse:
        """
        Iniciar processo de importação
        """
        logger.info(f"Iniciando importação: module={request.module}, filename={filename}")

        # Criar log
        import_log = await self.repository.create_import_log({
            "module": request.module,
            "format": request.format,
            "filename": filename,
            "status": ImportStatus.VALIDATING,
            "user_id": user_id,
            "started_at": datetime.utcnow(),
            "mapping": json.dumps(request.mapping)
        })

        try:
            # Validar
            validation_errors = await self.validate_import(
                file,
                request.format,
                request.module,
                request.mapping
            )

            if validation_errors and not request.skip_errors:
                await self.repository.update_import_log(import_log.id, {
                    "status": ImportStatus.FAILED,
                    "validation_errors": json.dumps([e.model_dump() for e in validation_errors]),
                    "completed_at": datetime.utcnow()
                })
                raise BusinessException(
                    f"Arquivo possui {len(validation_errors)} erros de validação"
                )

            # Atualizar status
            await self.repository.update_import_log(import_log.id, {
                "status": ImportStatus.VALIDATED
            })

            # Ler dados novamente
            file.seek(0)
            if request.format == ImportFormat.CSV:
                data = self._read_csv(file)
            elif request.format == ImportFormat.EXCEL:
                data = self._read_excel(file)
            elif request.format == ImportFormat.JSON:
                data = self._read_json(file)
            else:
                raise BusinessException(f"Formato não suportado: {request.format}")

            # Atualizar total de linhas
            await self.repository.update_import_log(import_log.id, {
                "total_rows": len(data),
                "status": ImportStatus.IMPORTING
            })

            # Processar importação
            if not request.dry_run:
                result = await self._process_import(
                    data,
                    request.module,
                    request.mapping,
                    import_log.id,
                    request.skip_errors
                )
            else:
                result = {"success": len(data), "failed": 0, "ids": []}

            # Atualizar log final
            await self.repository.update_import_log(import_log.id, {
                "status": ImportStatus.COMPLETED,
                "processed_rows": len(data),
                "success_rows": result["success"],
                "failed_rows": result["failed"],
                "rollback_data": json.dumps(result.get("ids", [])),
                "completed_at": datetime.utcnow()
            })

            logger.info(
                f"Importação concluída: {result['success']} sucesso, "
                f"{result['failed']} falhas"
            )

        except Exception as e:
            logger.error(f"Erro na importação: {str(e)}")
            await self.repository.update_import_log(import_log.id, {
                "status": ImportStatus.FAILED,
                "import_errors": str(e),
                "completed_at": datetime.utcnow()
            })
            raise

        # Retornar status
        import_log = await self.repository.get_import_log(import_log.id)
        return ImportStatusResponse.model_validate(import_log)

    async def _process_import(
        self,
        data: List[Dict[str, Any]],
        module: str,
        mapping: Dict[str, str],
        import_id: int,
        skip_errors: bool = False
    ) -> Dict[str, Any]:
        """
        Processar a importação de fato
        """
        success_count = 0
        failed_count = 0
        created_ids = []

        for idx, row in enumerate(data, start=1):
            try:
                # Aplicar mapeamento
                mapped_row = {}
                for file_col, system_field in mapping.items():
                    if file_col in row:
                        mapped_row[system_field] = row[file_col]

                # Importar registro
                created_id = await self._import_record(module, mapped_row)

                if created_id:
                    success_count += 1
                    created_ids.append(created_id)
                else:
                    failed_count += 1

                # Atualizar progresso
                if idx % 10 == 0:
                    await self.repository.update_import_log(import_id, {
                        "processed_rows": idx,
                        "success_rows": success_count,
                        "failed_rows": failed_count
                    })

            except Exception as e:
                logger.error(f"Erro ao importar linha {idx}: {str(e)}")
                failed_count += 1

                if not skip_errors:
                    raise

        return {
            "success": success_count,
            "failed": failed_count,
            "ids": created_ids
        }

    async def _import_record(
        self,
        module: str,
        data: Dict[str, Any]
    ) -> Optional[int]:
        """
        Importar um registro específico
        """
        # Aqui seria feita a importação real para o módulo específico
        # Por ora, apenas simulação

        if module == "produtos":
            # from app.modules.produtos.service import ProdutoService
            # service = ProdutoService(self.db)
            # produto = await service.criar_produto(data)
            # return produto.id
            pass
        elif module == "clientes":
            # from app.modules.clientes.service import ClienteService
            # service = ClienteService(self.db)
            # cliente = await service.criar_cliente(data)
            # return cliente.id
            pass

        # Placeholder
        return 1

    # ========================================================================
    # IMPORT - STATUS E ROLLBACK
    # ========================================================================

    async def get_import_status(self, import_id: int) -> ImportStatusResponse:
        """Obter status da importação"""
        import_log = await self.repository.get_import_log(import_id)
        if not import_log:
            raise NotFoundException(f"Importação {import_id} não encontrada")
        return ImportStatusResponse.model_validate(import_log)

    async def rollback_import(
        self,
        import_id: int,
        reason: Optional[str] = None
    ) -> ImportStatusResponse:
        """
        Fazer rollback de uma importação
        """
        logger.info(f"Iniciando rollback da importação {import_id}")

        import_log = await self.repository.get_import_log(import_id)
        if not import_log:
            raise NotFoundException(f"Importação {import_id} não encontrada")

        if not import_log.can_rollback:
            raise BusinessException("Esta importação não pode ser revertida")

        if import_log.status == ImportStatus.ROLLED_BACK:
            raise BusinessException("Importação já foi revertida")

        try:
            # Recuperar IDs criados
            if import_log.rollback_data:
                created_ids = json.loads(import_log.rollback_data)

                # Deletar registros criados
                for record_id in created_ids:
                    await self._delete_imported_record(import_log.module, record_id)

            # Atualizar status
            await self.repository.update_import_log(import_id, {
                "status": ImportStatus.ROLLED_BACK,
                "import_errors": f"Revertido: {reason}" if reason else "Revertido"
            })

            logger.info(f"Rollback concluído: {len(created_ids)} registros removidos")

        except Exception as e:
            logger.error(f"Erro no rollback: {str(e)}")
            raise BusinessException(f"Erro ao reverter importação: {str(e)}")

        import_log = await self.repository.get_import_log(import_id)
        return ImportStatusResponse.model_validate(import_log)

    async def _delete_imported_record(self, module: str, record_id: int):
        """Deletar registro importado"""
        # Implementar deleção por módulo
        pass

    # ========================================================================
    # EXPORT
    # ========================================================================

    async def start_export(
        self,
        request: ExportRequest,
        user_id: Optional[int] = None
    ) -> ExportStatusResponse:
        """
        Iniciar exportação
        """
        logger.info(f"Iniciando exportação: module={request.module}, format={request.format}")

        filename = request.filename or f"{request.module}_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.{request.format.value}"

        # Criar log
        export_log = await self.repository.create_export_log({
            "module": request.module,
            "format": request.format,
            "filename": filename,
            "status": ExportStatus.PROCESSING,
            "user_id": user_id,
            "started_at": datetime.utcnow(),
            "filters": json.dumps(request.filters) if request.filters else None
        })

        try:
            # Buscar dados
            data = await self._fetch_export_data(
                request.module,
                request.filters,
                request.columns
            )

            # Atualizar total
            await self.repository.update_export_log(export_log.id, {
                "total_records": len(data)
            })

            # Gerar arquivo
            file_path = await self._generate_export_file(
                data,
                request.format,
                filename,
                request.include_headers
            )

            # Atualizar log
            await self.repository.update_export_log(export_log.id, {
                "status": ExportStatus.COMPLETED,
                "file_path": file_path,
                "completed_at": datetime.utcnow()
            })

            logger.info(f"Exportação concluída: {len(data)} registros")

        except Exception as e:
            logger.error(f"Erro na exportação: {str(e)}")
            await self.repository.update_export_log(export_log.id, {
                "status": ExportStatus.FAILED,
                "error_message": str(e),
                "completed_at": datetime.utcnow()
            })
            raise

        export_log = await self.repository.get_export_log(export_log.id)
        return ExportStatusResponse.model_validate(export_log)

    async def _fetch_export_data(
        self,
        module: str,
        filters: Optional[Dict[str, Any]],
        columns: Optional[List[str]]
    ) -> List[Dict[str, Any]]:
        """
        Buscar dados para exportação
        """
        # Implementar busca por módulo
        # Por ora, dados de exemplo
        return [
            {"id": 1, "name": "Item 1", "value": 100},
            {"id": 2, "name": "Item 2", "value": 200},
        ]

    async def _generate_export_file(
        self,
        data: List[Dict[str, Any]],
        format: ExportFormat,
        filename: str,
        include_headers: bool = True
    ) -> str:
        """
        Gerar arquivo de exportação
        """
        # Criar diretório temporário
        temp_dir = tempfile.gettempdir()
        file_path = os.path.join(temp_dir, filename)

        if format == ExportFormat.CSV:
            self._write_csv(data, file_path, include_headers)
        elif format == ExportFormat.EXCEL:
            self._write_excel(data, file_path, include_headers)
        elif format == ExportFormat.JSON:
            self._write_json(data, file_path)
        else:
            raise BusinessException(f"Formato não suportado: {format}")

        return file_path

    # ========================================================================
    # HELPERS - LEITURA
    # ========================================================================

    def _read_csv(self, file: BinaryIO) -> List[Dict[str, Any]]:
        """Ler arquivo CSV"""
        content = file.read().decode('utf-8')
        reader = csv.DictReader(io.StringIO(content))
        return list(reader)

    def _read_excel(self, file: BinaryIO) -> List[Dict[str, Any]]:
        """Ler arquivo Excel"""
        wb = load_workbook(file)
        ws = wb.active

        # Primeira linha = headers
        headers = [cell.value for cell in ws[1]]

        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            data.append(row_dict)

        return data

    def _read_json(self, file: BinaryIO) -> List[Dict[str, Any]]:
        """Ler arquivo JSON"""
        content = file.read().decode('utf-8')
        data = json.loads(content)

        # Suportar tanto array quanto objeto único
        if isinstance(data, dict):
            return [data]
        return data

    # ========================================================================
    # HELPERS - ESCRITA
    # ========================================================================

    def _write_csv(
        self,
        data: List[Dict[str, Any]],
        file_path: str,
        include_headers: bool = True
    ):
        """Escrever arquivo CSV"""
        if not data:
            return

        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=data[0].keys())
            if include_headers:
                writer.writeheader()
            writer.writerows(data)

    def _write_excel(
        self,
        data: List[Dict[str, Any]],
        file_path: str,
        include_headers: bool = True
    ):
        """Escrever arquivo Excel com formatação"""
        wb = Workbook()
        ws = wb.active

        if not data:
            wb.save(file_path)
            return

        # Headers
        headers = list(data[0].keys())
        if include_headers:
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

        # Data
        start_row = 2 if include_headers else 1
        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))

        wb.save(file_path)

    def _write_json(self, data: List[Dict[str, Any]], file_path: str):
        """Escrever arquivo JSON"""
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False, default=str)

    # ========================================================================
    # HELPERS - VALIDAÇÃO E MAPEAMENTO
    # ========================================================================

    def _suggest_column_mapping(
        self,
        columns: List[str],
        module: str
    ) -> Dict[str, str]:
        """
        Sugerir mapeamento de colunas
        """
        # Mapeamentos comuns por módulo
        mappings = {
            "produtos": {
                "codigo": ["codigo", "code", "sku", "cod"],
                "descricao": ["descricao", "description", "desc", "nome", "name"],
                "preco_venda": ["preco", "price", "preco_venda", "valor"],
                "preco_custo": ["custo", "cost", "preco_custo"],
                "estoque_minimo": ["estoque_min", "min_stock", "estoque_minimo"],
            },
            "clientes": {
                "nome": ["nome", "name", "razao_social"],
                "cpf_cnpj": ["cpf", "cnpj", "documento", "cpf_cnpj"],
                "email": ["email", "e-mail"],
                "telefone": ["telefone", "phone", "tel", "fone"],
            }
        }

        suggestions = {}
        module_mappings = mappings.get(module, {})

        for col in columns:
            col_lower = col.lower().strip()
            for field, aliases in module_mappings.items():
                if col_lower in [alias.lower() for alias in aliases]:
                    suggestions[col] = field
                    break

        return suggestions

    def _get_required_fields(self, module: str) -> List[str]:
        """Obter campos obrigatórios por módulo"""
        required = {
            "produtos": ["codigo", "descricao", "preco_venda"],
            "clientes": ["nome", "cpf_cnpj"],
            "fornecedores": ["razao_social", "cnpj"],
        }
        return required.get(module, [])

    async def _validate_module_data(
        self,
        module: str,
        data: Dict[str, Any],
        row_number: int
    ) -> List[ImportValidationError]:
        """
        Validações específicas por módulo
        """
        errors = []

        # Validações por módulo
        if module == "produtos":
            # Validar preço
            if "preco_venda" in data:
                try:
                    preco = float(data["preco_venda"])
                    if preco <= 0:
                        errors.append(ImportValidationError(
                            row=row_number,
                            column="preco_venda",
                            value=data["preco_venda"],
                            error="Preço deve ser maior que zero"
                        ))
                except (ValueError, TypeError):
                    errors.append(ImportValidationError(
                        row=row_number,
                        column="preco_venda",
                        value=data["preco_venda"],
                        error="Preço inválido"
                    ))

        elif module == "clientes":
            # Validar CPF/CNPJ
            if "cpf_cnpj" in data:
                doc = str(data["cpf_cnpj"]).replace(".", "").replace("-", "").replace("/", "")
                if len(doc) not in [11, 14]:
                    errors.append(ImportValidationError(
                        row=row_number,
                        column="cpf_cnpj",
                        value=data["cpf_cnpj"],
                        error="CPF/CNPJ inválido"
                    ))

        return errors

    # ========================================================================
    # TEMPLATES
    # ========================================================================

    async def create_template(
        self,
        data: ImportTemplateCreate,
        user_id: Optional[int] = None
    ) -> ImportTemplateResponse:
        """Criar template de importação"""
        # Verificar se já existe
        existing = await self.repository.get_template_by_name(data.name, data.module)
        if existing:
            raise BusinessException(
                f"Template '{data.name}' já existe para o módulo '{data.module}'"
            )

        template = await self.repository.create_template({
            "name": data.name,
            "module": data.module,
            "description": data.description,
            "format": data.format,
            "column_mapping": json.dumps(data.column_mapping),
            "required_columns": json.dumps(data.required_columns),
            "validation_rules": json.dumps(data.validation_rules) if data.validation_rules else None,
            "created_by": user_id
        })

        return ImportTemplateResponse.model_validate(template)

    async def list_templates(
        self,
        module: Optional[str] = None,
        skip: int = 0,
        limit: int = 100
    ) -> List[ImportTemplateResponse]:
        """Listar templates"""
        templates = await self.repository.list_templates(module, True, skip, limit)
        return [ImportTemplateResponse.model_validate(t) for t in templates]

    # ========================================================================
    # ESTATÍSTICAS
    # ========================================================================

    async def get_statistics(self) -> ImportExportStats:
        """Obter estatísticas gerais"""
        import_stats = await self.repository.get_import_stats()
        export_stats = await self.repository.get_export_stats()

        most_imported = await self.repository.get_most_imported_module()
        most_exported = await self.repository.get_most_exported_module()

        return ImportExportStats(
            total_imports=import_stats["total"],
            successful_imports=import_stats["successful"],
            failed_imports=import_stats["failed"],
            total_exports=export_stats["total"],
            successful_exports=export_stats["successful"],
            failed_exports=export_stats["failed"],
            most_imported_module=most_imported,
            most_exported_module=most_exported
        )
