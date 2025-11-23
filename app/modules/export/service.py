"""
Service para exportação de dados em Excel e CSV
"""

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, cast, Date
from datetime import datetime, date, timedelta
from typing import List, Dict, Any, Optional
import io
import csv

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from app.modules.vendas.models import Venda, ItemVenda
from app.modules.orcamentos.models import Orcamento, ItemOrcamento
from app.modules.produtos.models import Produto
from app.modules.categorias.models import Categoria
from app.modules.clientes.models import Cliente
from app.modules.auth.models import User
from .schemas import ExportFiltros


class ExportService:
    """Service para gerar exports em Excel e CSV"""

    def __init__(self, db: AsyncSession):
        self.db = db

    async def export_dashboard_stats(
        self, formato: str, tipo: str, filtros: Optional[ExportFiltros] = None
    ) -> io.BytesIO:
        """
        Exportar dados do dashboard

        Args:
            formato: 'excel' ou 'csv'
            tipo: Tipo de dados (stats, vendas_dia, produtos, vendedores, status)
            filtros: Filtros opcionais
        """
        if tipo == "vendas_dia":
            data = await self._get_vendas_por_dia(filtros)
            headers = ["Data", "Vendas", "Faturamento"]
            rows = [[d["data"], d["vendas"], f"R$ {d['faturamento']:.2f}"] for d in data]
            title = "Vendas por Dia"

        elif tipo == "produtos":
            data = await self._get_produtos_mais_vendidos(filtros)
            headers = ["Produto", "Quantidade", "Faturamento"]
            rows = [[d["produto"], d["quantidade"], f"R$ {d['faturamento']:.2f}"] for d in data]
            title = "Produtos Mais Vendidos"

        elif tipo == "vendedores":
            data = await self._get_vendas_por_vendedor(filtros)
            headers = ["Vendedor", "Vendas", "Faturamento", "Ticket Médio"]
            rows = [
                [d["vendedor"], d["vendas"], f"R$ {d['faturamento']:.2f}", f"R$ {d['ticket_medio']:.2f}"]
                for d in data
            ]
            title = "Vendas por Vendedor"

        elif tipo == "status":
            data = await self._get_vendas_por_status(filtros)
            headers = ["Status", "Quantidade", "Percentual"]
            rows = [[d["status"], d["quantidade"], f"{d['percentual']:.1f}%"] for d in data]
            title = "Vendas por Status"

        else:  # stats
            data = await self._get_dashboard_stats(filtros)
            headers = ["Métrica", "Valor"]
            rows = [
                ["Vendas Hoje", str(data.get("vendas_hoje", 0))],
                ["Vendas do Mês", str(data.get("vendas_mes", 0))],
                ["Faturamento do Mês", f"R$ {data.get('faturamento_mes', 0):.2f}"],
                ["Ticket Médio", f"R$ {data.get('ticket_medio', 0):.2f}"],
                ["Pedidos Abertos", str(data.get("pedidos_abertos", 0))],
                ["Crescimento (%)", f"{data.get('crescimento_mes', 0):.1f}%"],
            ]
            title = "Estatísticas do Dashboard"

        if formato == "excel":
            return self._create_excel(title, headers, rows)
        else:
            return self._create_csv(headers, rows)

    async def export_orcamentos(
        self,
        formato: str,
        filtros: Optional[ExportFiltros] = None,
        ids: Optional[List[int]] = None
    ) -> io.BytesIO:
        """Exportar orçamentos"""
        query = select(Orcamento).join(Cliente, Orcamento.cliente_id == Cliente.id)

        if ids:
            query = query.where(Orcamento.id.in_(ids))
        else:
            if filtros:
                if filtros.data_inicio:
                    query = query.where(cast(Orcamento.data_orcamento, Date) >= filtros.data_inicio)
                if filtros.data_fim:
                    query = query.where(cast(Orcamento.data_orcamento, Date) <= filtros.data_fim)
                if filtros.cliente_id:
                    query = query.where(Orcamento.cliente_id == filtros.cliente_id)
                if filtros.status:
                    query = query.where(Orcamento.status == filtros.status)

        result = await self.db.execute(query)
        orcamentos = result.scalars().all()

        headers = [
            "ID", "Data", "Cliente", "Vendedor", "Status",
            "Valor Total", "Desconto", "Validade"
        ]

        rows = []
        for orc in orcamentos:
            cliente = await self.db.get(Cliente, orc.cliente_id)
            vendedor = await self.db.get(User, orc.vendedor_id) if orc.vendedor_id else None

            rows.append([
                str(orc.id),
                orc.data_orcamento.strftime("%d/%m/%Y"),
                cliente.nome if cliente else "N/A",
                vendedor.nome if vendedor else "N/A",
                orc.status,
                f"R$ {float(orc.valor_total):.2f}",
                f"R$ {float(orc.desconto or 0):.2f}",
                orc.validade.strftime("%d/%m/%Y") if orc.validade else "N/A"
            ])

        if formato == "excel":
            return self._create_excel("Orçamentos", headers, rows)
        else:
            return self._create_csv(headers, rows)

    async def export_vendas(
        self,
        formato: str,
        filtros: Optional[ExportFiltros] = None,
        ids: Optional[List[int]] = None
    ) -> io.BytesIO:
        """Exportar vendas"""
        query = select(Venda).join(Cliente, Venda.cliente_id == Cliente.id)

        if ids:
            query = query.where(Venda.id.in_(ids))
        else:
            if filtros:
                if filtros.data_inicio:
                    query = query.where(cast(Venda.data_venda, Date) >= filtros.data_inicio)
                if filtros.data_fim:
                    query = query.where(cast(Venda.data_venda, Date) <= filtros.data_fim)
                if filtros.cliente_id:
                    query = query.where(Venda.cliente_id == filtros.cliente_id)
                if filtros.vendedor_id:
                    query = query.where(Venda.vendedor_id == filtros.vendedor_id)
                if filtros.status:
                    query = query.where(Venda.status == filtros.status)

        result = await self.db.execute(query)
        vendas = result.scalars().all()

        headers = [
            "ID", "Data", "Cliente", "Vendedor", "Status",
            "Forma Pagamento", "Valor Total", "Desconto", "Valor Final"
        ]

        rows = []
        for venda in vendas:
            cliente = await self.db.get(Cliente, venda.cliente_id)
            vendedor = await self.db.get(User, venda.vendedor_id) if venda.vendedor_id else None

            rows.append([
                str(venda.id),
                venda.data_venda.strftime("%d/%m/%Y %H:%M"),
                cliente.nome if cliente else "N/A",
                vendedor.nome if vendedor else "N/A",
                venda.status,
                venda.forma_pagamento or "N/A",
                f"R$ {float(venda.valor_total):.2f}",
                f"R$ {float(venda.desconto or 0):.2f}",
                f"R$ {float(venda.valor_final):.2f}"
            ])

        if formato == "excel":
            return self._create_excel("Vendas", headers, rows)
        else:
            return self._create_csv(headers, rows)

    async def export_produtos(
        self,
        formato: str,
        categoria_id: Optional[int] = None,
        apenas_ativos: bool = True
    ) -> io.BytesIO:
        """Exportar produtos"""
        query = select(Produto).join(Categoria, Produto.categoria_id == Categoria.id)

        if categoria_id:
            query = query.where(Produto.categoria_id == categoria_id)
        if apenas_ativos:
            query = query.where(Produto.ativo == True)

        result = await self.db.execute(query)
        produtos = result.scalars().all()

        headers = [
            "Código", "Descrição", "Categoria", "Unidade",
            "Preço Custo", "Preço Venda", "Estoque", "Ativo"
        ]

        rows = []
        for produto in produtos:
            categoria = await self.db.get(Categoria, produto.categoria_id)

            rows.append([
                produto.codigo,
                produto.descricao,
                categoria.nome if categoria else "N/A",
                produto.unidade,
                f"R$ {float(produto.preco_custo or 0):.2f}",
                f"R$ {float(produto.preco_venda):.2f}",
                str(float(produto.estoque_atual or 0)),
                "Sim" if produto.ativo else "Não"
            ])

        if formato == "excel":
            return self._create_excel("Produtos", headers, rows)
        else:
            return self._create_csv(headers, rows)

    # Helper methods para buscar dados
    async def _get_vendas_por_dia(self, filtros: Optional[ExportFiltros]) -> List[Dict]:
        """Buscar vendas agrupadas por dia"""
        hoje = date.today()
        data_inicio = filtros.data_inicio if filtros and filtros.data_inicio else hoje - timedelta(days=30)
        data_fim = filtros.data_fim if filtros and filtros.data_fim else hoje

        query = (
            select(
                cast(Venda.data_venda, Date).label("data"),
                func.count(Venda.id).label("vendas"),
                func.sum(Venda.valor_final).label("faturamento")
            )
            .where(
                and_(
                    cast(Venda.data_venda, Date) >= data_inicio,
                    cast(Venda.data_venda, Date) <= data_fim
                )
            )
            .group_by(cast(Venda.data_venda, Date))
            .order_by(cast(Venda.data_venda, Date))
        )

        result = await self.db.execute(query)
        rows = result.all()

        return [
            {
                "data": row.data.strftime("%d/%m/%Y"),
                "vendas": row.vendas,
                "faturamento": float(row.faturamento or 0)
            }
            for row in rows
        ]

    async def _get_produtos_mais_vendidos(self, filtros: Optional[ExportFiltros]) -> List[Dict]:
        """Buscar produtos mais vendidos"""
        query = (
            select(
                Produto.descricao,
                func.sum(ItemVenda.quantidade).label("quantidade"),
                func.sum(ItemVenda.total_item).label("faturamento")
            )
            .join(ItemVenda, Produto.id == ItemVenda.produto_id)
            .join(Venda, ItemVenda.venda_id == Venda.id)
        )

        if filtros:
            if filtros.data_inicio:
                query = query.where(cast(Venda.data_venda, Date) >= filtros.data_inicio)
            if filtros.data_fim:
                query = query.where(cast(Venda.data_venda, Date) <= filtros.data_fim)

        query = (
            query.group_by(Produto.id, Produto.descricao)
            .order_by(func.sum(ItemVenda.quantidade).desc())
            .limit(20)
        )

        result = await self.db.execute(query)
        rows = result.all()

        return [
            {
                "produto": row.descricao,
                "quantidade": float(row.quantidade or 0),
                "faturamento": float(row.faturamento or 0)
            }
            for row in rows
        ]

    async def _get_vendas_por_vendedor(self, filtros: Optional[ExportFiltros]) -> List[Dict]:
        """Buscar vendas por vendedor"""
        query = (
            select(
                User.nome,
                func.count(Venda.id).label("vendas"),
                func.sum(Venda.valor_final).label("faturamento")
            )
            .join(Venda, User.id == Venda.vendedor_id)
        )

        if filtros:
            if filtros.data_inicio:
                query = query.where(cast(Venda.data_venda, Date) >= filtros.data_inicio)
            if filtros.data_fim:
                query = query.where(cast(Venda.data_venda, Date) <= filtros.data_fim)

        query = query.group_by(User.id, User.nome).order_by(func.sum(Venda.valor_final).desc())

        result = await self.db.execute(query)
        rows = result.all()

        return [
            {
                "vendedor": row.nome,
                "vendas": row.vendas,
                "faturamento": float(row.faturamento or 0),
                "ticket_medio": float(row.faturamento or 0) / row.vendas if row.vendas > 0 else 0
            }
            for row in rows
        ]

    async def _get_vendas_por_status(self, filtros: Optional[ExportFiltros]) -> List[Dict]:
        """Buscar vendas por status"""
        query = select(
            Venda.status,
            func.count(Venda.id).label("quantidade")
        )

        if filtros:
            if filtros.data_inicio:
                query = query.where(cast(Venda.data_venda, Date) >= filtros.data_inicio)
            if filtros.data_fim:
                query = query.where(cast(Venda.data_venda, Date) <= filtros.data_fim)

        query = query.group_by(Venda.status)

        result = await self.db.execute(query)
        rows = result.all()

        total = sum(row.quantidade for row in rows)

        return [
            {
                "status": row.status,
                "quantidade": row.quantidade,
                "percentual": (row.quantidade / total * 100) if total > 0 else 0
            }
            for row in rows
        ]

    async def _get_dashboard_stats(self, filtros: Optional[ExportFiltros]) -> Dict:
        """Buscar estatísticas do dashboard"""
        hoje = date.today()
        inicio_mes = date(hoje.year, hoje.month, 1)

        # Vendas hoje
        query_hoje = select(func.count(Venda.id)).where(cast(Venda.data_venda, Date) == hoje)
        result = await self.db.execute(query_hoje)
        vendas_hoje = result.scalar() or 0

        # Vendas do mês
        query_mes = select(
            func.count(Venda.id).label("total"),
            func.sum(Venda.valor_final).label("faturamento")
        ).where(cast(Venda.data_venda, Date) >= inicio_mes)

        result = await self.db.execute(query_mes)
        row = result.first()
        vendas_mes = row.total if row else 0
        faturamento_mes = float(row.faturamento or 0) if row else 0

        # Ticket médio
        ticket_medio = faturamento_mes / vendas_mes if vendas_mes > 0 else 0

        # Pedidos abertos
        query_abertos = select(func.count(Venda.id)).where(Venda.status == "aberto")
        result = await self.db.execute(query_abertos)
        pedidos_abertos = result.scalar() or 0

        return {
            "vendas_hoje": vendas_hoje,
            "vendas_mes": vendas_mes,
            "faturamento_mes": faturamento_mes,
            "ticket_medio": ticket_medio,
            "pedidos_abertos": pedidos_abertos,
            "crescimento_mes": 0  # Placeholder
        }

    def _create_excel(self, title: str, headers: List[str], rows: List[List[Any]]) -> io.BytesIO:
        """Criar arquivo Excel"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl não está instalado. Execute: pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel limit

        # Título
        ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
        title_cell = ws["A1"]
        title_cell.value = title
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")

        # Headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Data rows
        for row_idx, row_data in enumerate(rows, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal="left")

        # Auto-ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Salvar em BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _create_csv(self, headers: List[str], rows: List[List[Any]]) -> io.BytesIO:
        """Criar arquivo CSV"""
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)

        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)

        # Converter para BytesIO
        bytes_output = io.BytesIO()
        bytes_output.write(output.getvalue().encode('utf-8-sig'))  # BOM para Excel
        bytes_output.seek(0)
        return bytes_output
